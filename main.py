from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

data = {
	"Scofield": {
		"Music": 65,
		"Science": 68,
		"Geometry": 19,
		"Chemistry": 89
	},
	"Lincoln": {
		"Music": 55,
		"Science": 22,
		"Geometry": 96,
		"Chemistry": 95
	},
	"Julia": {
		"Music": 100,
		"Science": 52,
		"Geometry": 75,
		"Chemistry": 92
	},
	"Nicole": {
		"Music": 30,
		"Science": 70,
		"Geometry": 33,
		"Chemistry": 100
	},
	"Rose": {
		"Music": 100,
		"Science": 46,
		"Geometry": 80,
		"Chemistry": 60
	}
}

wb = Workbook()
ws = wb.active
ws.title = "School Grades"

headings = ['Name'] + list(data['Scofield'].keys())
ws.append(headings)

for person in data:
	grades = list(data[person].values())
	ws.append([person] + grades)

for col in range(2, len(data['Scofield']) + 2):
	char = get_column_letter(col)
	ws[char + "7"] = f"=SUM({char + '2'}:{char + '6'})/{len(data)}"

for col in range(1, 6):
	ws[get_column_letter(col) + '1'].font = Font(bold=True, color="0000CCFF")
ws[get_column_letter(1) + "7"] = "AVG"
wb.save("TurtleCode.xlsx")